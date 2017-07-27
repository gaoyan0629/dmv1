import sys
import os
import xlrd
import openpyxl
import re
from util import *

class ExcelDoc:
	def __init__(self,fileName,tabName='DD Data',headerLoc=0,mode = 'r'):
		self.fileName = fileName
		self.mode = mode.lower()
		self.tabName = tabName
		suffix = re.findall(r"""(\.xls[x]?$)""",self.fileName)
		if len(suffix) != 1:
			raise fileSuffixError("Error","Suffix of file of {} is wrong".format(fileName))
		self.fileSuffix =suffix[0] 
		if (self.fileSuffix not in [".xls",".xlsx"]):
			raise fileSuffixError("Error","Suffix of file of {} is wrong".format(fileName))
		self.headerLoc = headerLoc
		self.dataLoc = headerLoc + 1
		self.header = []
		self.data = []
		self.rowCount = 0
		self.colCount = 0
		if self.mode == 'r':
			if self.fileSuffix == '.xls': 
				self.workBook = xlrd.open_workbook(self.fileName)
				self.workSheet = self.workBook.sheet_by_name(self.tabName)
				self.maxRow = self.workSheet.nrows
				self.maxColumn = self.workSheet.ncols
				self.headStart = self.headerLoc - 1
				self.rowStart = self.headerLoc
				self.rowEnd = self.maxRow
				self.columnStart = 0
				self.columnEnd = self.maxColumn
			elif self.fileSuffix =='.xlsx':
				self.workBook = openpyxl.load_workbook(self.fileName)
				self.workSheet = self.workBook.get_sheet_by_name(self.tabName)
				self.maxRow = self.workSheet.max_row
				self.maxColumn = self.workSheet.max_column
				self.headStart = self.headerLoc 
				self.rowStart = self.headerLoc+1
				self.rowEnd = self.maxRow + 1
				self.columnStart = 1
				self.columnEnd = self.maxColumn +1
			self.__read__()
		elif self.mode == 'w':
			if self.fileSuffix =='.xlsx': #for xls need xlwt
				self.workBook = openpyxl.Workbook()
				self.workSheet = self.workBook.create_sheet(index=0, 							title=self.tabName)

	def __read__(self):
		self.rowCount = self.maxRow - self.headerLoc
		self.colCount = self.maxColumn
		if self.headerLoc > 0:
			for col in range(self.columnStart,self.columnEnd):
				if self.fileSuffix =='.xls':
					self.header.append(self.workSheet.cell(self.headStart,col).value)
				else:
					self.header.append(self.workSheet.cell(row = self.headStart,column = col).value)
		for row in range(self.rowStart,self.rowEnd):
			DataRow = []
			for col in range(self.columnStart,self.columnEnd):
				if self.fileSuffix =='.xls':
            				DataRow.append(self.workSheet.cell(row, col).value)
				else:
            				DataRow.append(self.workSheet.cell(row=row, column=col).value)
			self.data.append(DataRow)
		return True

	def write(self,data):
		if self.fileSuffix == '.xls': 
			raise NoSupportError("writing to xls is still not in support")
			
		elif self.fileSuffix == '.xlsx':
			for row in range(len(data)):
				for col in range(len(data[row])):
					if not data[row][col]:
						continue
					self.workSheet.cell(row = row+1,column = col +1).value = data[row][col] 
		self.workBook.save(self.fileName)

	def __iter__(self):
		self.count = 0
		self.countMax = len(self.data)
		return self
	def next(self):
		if self.count < self.countMax:
			self.count += 1
			return self.data[self.count - 1]
		else:
			raise StopIteration

	def getDataOnLoc(self,headLoc):
		ret =[]
		for i in headLoc:
			ret.append(zip(*self.data)[i])
		return list(zip(*ret))

	def __del__(self):
		pass


if __name__ == "__main__":
	#myFile = ["DD_101_ABSP.xls","DD_126_NTSP.xls", "DD_101_ABSP1.xlsx"]
	#myFile = ["DD_126_NTSP.xls", "DD_101_ABSP1.xlsx"]
	myFile = ["DD_126_NTSP.xls"]
	for myFileName in myFile:
		reader = ExcelDoc(fileName = myFileName,headerLoc = 2) 
		for i in reader:
			print i
		writer = ExcelDoc(fileName = "DD_101_ABSP.xlsx", mode= "w")
		writer.write([reader.header] + reader.data)
